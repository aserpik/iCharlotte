"""
Update Excel case list with data from iCharlotte master_cases.db.

Maps:
  - Excel "File No." <-> DB "file_number"
  - Excel "FM/HP" <-> DB "assigned_attorney"
  - Excel "Trial Date" <-> DB "trial_date"

Only updates cells where DB has a non-empty value that differs from Excel.
Highlights changes in yellow and adds a "Changes" column.
"""

import sqlite3
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, date
from collections import OrderedDict

# === Paths ===
EXCEL_PATH = r"C:\Users\ASerpik.DESKTOP-MRIMK0D\OneDrive - Bordin Semmer LLP\Desktop\AS Team Case List.xlsx"
DB_PATH = r"C:\geminiterminal2\.gemini\case_data\master_cases.db"
SHEET_NAME = "AS TEAM CASE LIST"

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def normalize_file_number(val):
    """
    Normalize a file number to a canonical string form.
    
    Excel stores file numbers as either:
      - float (e.g., 5800.056) -> str gives "5800.056"
      - str (e.g., "5800.060") -> for trailing-zero cases
    
    DB stores as str (e.g., "5800.060").
    
    Strategy: Format with 3 decimal places to match DB convention.
    """
    if val is None:
        return None
    
    s = str(val).strip()
    if not s:
        return None
    
    try:
        f = float(s)
        return f"{f:.3f}"
    except ValueError:
        return s


def parse_db_date(date_str):
    """Parse a DB date string (YYYY-MM-DD) to a datetime object."""
    if not date_str or not date_str.strip():
        return None
    try:
        return datetime.strptime(date_str.strip(), "%Y-%m-%d")
    except ValueError:
        return None


def excel_date_to_date(val):
    """Extract a date from an Excel cell value (could be datetime or None)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str) and val.strip():
        try:
            return datetime.strptime(val.strip(), "%Y-%m-%d").date()
        except ValueError:
            pass
    return None


def format_date_for_display(val):
    """Format a date value for display in change log."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return str(val)


def main():
    # --- Load DB data ---
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT file_number, assigned_attorney, trial_date FROM cases")
    
    db_data = {}
    for file_number, assigned_attorney, trial_date in cur.fetchall():
        fn_norm = normalize_file_number(file_number)
        if fn_norm:
            db_data[fn_norm] = {
                "assigned_attorney": (assigned_attorney or "").strip(),
                "trial_date": (trial_date or "").strip(),
            }
    conn.close()
    print(f"Loaded {len(db_data)} records from DB.")

    # --- Load Excel ---
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    
    # --- Identify columns by header (row 1) ---
    col_map = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header is not None:
            header_stripped = header.strip()
            if header_stripped == "File No.":
                col_map["file_no"] = col
            elif header_stripped == "FM/HP":
                col_map["fmhp"] = col
            elif header_stripped == "Trial Date":
                col_map["trial_date"] = col
    
    print(f"Column mapping: {col_map}")
    
    if "file_no" not in col_map:
        print("ERROR: Could not find 'File No.' column!")
        return
    if "fmhp" not in col_map:
        print("ERROR: Could not find 'FM/HP' column!")
        return
    if "trial_date" not in col_map:
        print("ERROR: Could not find 'Trial Date' column!")
        return
    
    # --- Add "Changes" column ---
    changes_col = ws.max_column + 1
    ws.cell(row=1, column=changes_col, value="Changes")
    
    # --- Iterate rows and update ---
    total_changes = 0
    all_changes = OrderedDict()
    
    for row in range(2, ws.max_row + 1):
        raw_file_no = ws.cell(row=row, column=col_map["file_no"]).value
        fn_norm = normalize_file_number(raw_file_no)
        
        if fn_norm is None:
            continue
        
        if fn_norm not in db_data:
            continue
        
        db_record = db_data[fn_norm]
        row_changes = []
        
        # --- Compare FM/HP ---
        db_attorney = db_record["assigned_attorney"]
        if db_attorney:  # Only update if DB has a non-empty value
            excel_fmhp = ws.cell(row=row, column=col_map["fmhp"]).value
            excel_fmhp_str = (str(excel_fmhp).strip() if excel_fmhp is not None else "")
            
            if excel_fmhp_str != db_attorney:
                old_val = excel_fmhp_str if excel_fmhp_str else ""
                ws.cell(row=row, column=col_map["fmhp"]).value = db_attorney
                ws.cell(row=row, column=col_map["fmhp"]).fill = YELLOW_FILL
                row_changes.append(f"FM/HP: {old_val}\u2192{db_attorney}")
                total_changes += 1
        
        # --- Compare Trial Date ---
        db_trial_str = db_record["trial_date"]
        if db_trial_str:  # Only update if DB has a non-empty value
            db_trial_dt = parse_db_date(db_trial_str)
            if db_trial_dt is not None and db_trial_dt.date() >= date.today():
                excel_trial_val = ws.cell(row=row, column=col_map["trial_date"]).value
                excel_trial_date = excel_date_to_date(excel_trial_val)
                db_trial_date = db_trial_dt.date()
                
                if excel_trial_date != db_trial_date:
                    old_val = format_date_for_display(excel_trial_val) if excel_trial_val else ""
                    new_val = db_trial_dt.strftime("%Y-%m-%d")
                    # Write as datetime so Excel displays it properly
                    ws.cell(row=row, column=col_map["trial_date"]).value = db_trial_dt
                    ws.cell(row=row, column=col_map["trial_date"]).fill = YELLOW_FILL
                    row_changes.append(f"Trial Date: {old_val}\u2192{new_val}")
                    total_changes += 1
        
        # --- Record changes ---
        if row_changes:
            change_text = "; ".join(row_changes)
            ws.cell(row=row, column=changes_col).value = change_text
            all_changes[fn_norm] = row_changes
    
    # --- Save ---
    wb.save(EXCEL_PATH)
    print(f"\nSaved to: {EXCEL_PATH}")
    
    # --- Print summary ---
    print("\n" + "=" * 60)
    print("CHANGE SUMMARY")
    print("=" * 60)
    
    if not all_changes:
        print("No changes were made.")
    else:
        for fn, changes in all_changes.items():
            print(f"\n  {fn}:")
            for c in changes:
                print(f"    - {c}")
    
    print(f"\n{'=' * 60}")
    print(f"TOTAL CHANGES: {total_changes}")
    print(f"FILES AFFECTED: {len(all_changes)}")
    print("=" * 60)


if __name__ == "__main__":
    main()
