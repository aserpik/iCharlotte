"""
Verify trial dates from case list spreadsheet by running docket agents.
Also fixes date formatting to "Month DD, YYYY".
"""
import os
import sys
import json
import subprocess
import time
import datetime
import openpyxl
from openpyxl.styles import numbers

XLSX_PATH = r"C:\Users\ASerpik.DESKTOP-MRIMK0D\OneDrive - Bordin Semmer LLP\Desktop\AS Team Case List - Copy.xlsx"

# Jurisdiction mapping from spreadsheet values to venue_county values used by docket.py
SUPPORTED_JURISDICTIONS = {
    "los angeles": "los angeles",
    "san bernardino": "san bernardino",
    "riverside": "riverside",
    "san diego": "san diego",
    "orange": "orange",
    "kern": "kern",
    "sacramento": "sacramento",
    "santa clara": "santa clara",
    "alameda": "alameda",
    "mariposa": "mariposa",
    "yolo": "yolo",
    "lake": "lake",
    "fresno": "fresno",  # Check if supported - no scraper found
    "placer": "placer",  # Check if supported - no scraper found
}

# These jurisdictions have no scraper
UNSUPPORTED = {"southern district", "central district", "fresno", "placer"}


def format_date_str(dt):
    """Format datetime to 'Month DD, YYYY'"""
    if isinstance(dt, datetime.datetime):
        return dt.strftime("%B %d, %Y").replace(" 0", " ")
    if isinstance(dt, str):
        # Try parsing YYYY-MM-DD
        try:
            d = datetime.datetime.strptime(dt.strip(), "%Y-%m-%d")
            return d.strftime("%B %d, %Y").replace(" 0", " ")
        except ValueError:
            pass
    return str(dt) if dt else ""


def is_red_row(row):
    """Check if any cell in the row has a red fill."""
    for c in row:
        fg = c.fill.fgColor
        if fg and fg.rgb and isinstance(fg.rgb, str):
            rgb_upper = fg.rgb.upper()
            if 'FF0000' in rgb_upper or rgb_upper == 'FFFF0000':
                return True
    return False


def main():
    print("=" * 70)
    print("TRIAL DATE VERIFICATION AND FORMATTING TOOL")
    print("=" * 70)

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    # Collect cases
    litigation_cases = []  # (row_idx, file_no, case_name, case_no, jurisdiction, trial_date)
    all_trial_date_rows = []  # (row_idx, file_no, trial_date)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        file_no = row[0].value
        if file_no is None:
            continue

        file_no_str = str(file_no).strip()
        case_name = row[1].value or ""
        fm_hp = row[2].value or ""
        pre_lit = str(row[3].value or "").strip().lower()
        case_no = str(row[4].value or "").strip()
        jurisdiction = str(row[5].value or "").strip()
        trial_date = row[6].value

        row_idx = row[0].row

        # Task 2: Track all rows with trial dates for formatting fix
        if trial_date is not None:
            all_trial_date_rows.append((row_idx, file_no_str, trial_date))

        # Task 1: Litigation cases not highlighted red
        is_red = is_red_row(row)
        is_litigation = 'lit' in pre_lit and 'pre' not in pre_lit

        if not is_red and is_litigation:
            litigation_cases.append((row_idx, file_no_str, case_name, case_no, jurisdiction, trial_date))

    # ============================================================
    # TASK 2: Fix date formatting for ALL trial dates
    # ============================================================
    print(f"\n--- TASK 2: Fixing date format for {len(all_trial_date_rows)} rows with trial dates ---")

    for row_idx, file_no, trial_date in all_trial_date_rows:
        cell = ws.cell(row=row_idx, column=7)
        formatted = format_date_str(trial_date)
        if formatted:
            cell.value = formatted
            cell.number_format = '@'  # Text format
            print(f"  {file_no}: {trial_date} -> {formatted}")

    # Save formatting changes
    wb.save(XLSX_PATH)
    print(f"\nSaved date formatting changes to spreadsheet.")

    # ============================================================
    # TASK 1: Run docket agents for litigation cases
    # ============================================================
    print(f"\n--- TASK 1: Running docket agents for {len(litigation_cases)} litigation cases ---")

    # Separate into supported and unsupported jurisdictions
    to_run = []
    cannot_run = []

    for row_idx, file_no, case_name, case_no, jurisdiction, trial_date in litigation_cases:
        j_lower = jurisdiction.lower().strip()
        if not j_lower or j_lower in ('none', 'n/a', 'null', ''):
            cannot_run.append((file_no, case_name, "No jurisdiction listed"))
        elif j_lower in UNSUPPORTED:
            cannot_run.append((file_no, case_name, f"Unsupported jurisdiction: {jurisdiction}"))
        elif not case_no or case_no.lower() in ('none', 'n/a', 'null', ''):
            cannot_run.append((file_no, case_name, "No case number listed"))
        else:
            to_run.append((row_idx, file_no, case_name, case_no, jurisdiction, trial_date))

    print(f"\n  Will run docket agent for {len(to_run)} cases")
    print(f"  Cannot run for {len(cannot_run)} cases")

    # Run docket agents in batches using docket.py's built-in dispatcher
    # We'll run them 5 at a time
    file_numbers_to_run = [fn for _, fn, _, _, _, _ in to_run]

    if file_numbers_to_run:
        print(f"\n  Running docket agent for: {', '.join(file_numbers_to_run)}")

        # Use docket.py's multi-file support — pass comma-separated (parse_file_numbers splits on comma)
        csv_arg = ",".join(file_numbers_to_run)
        cmd = [sys.executable, os.path.join("Scripts", "docket.py"), csv_arg, "--headless"]
        print(f"  Command: {cmd[0]} {cmd[1]} <{len(file_numbers_to_run)} files comma-separated> --headless")

        try:
            result = subprocess.run(
                cmd,
                cwd=r"C:\geminiterminal2",
                timeout=3600,  # 1 hour max
                capture_output=True,
                text=True
            )
            print(f"  Docket agent completed with return code: {result.returncode}")
            if result.stdout:
                # Print last 50 lines of output
                lines = result.stdout.strip().split('\n')
                for line in lines[-50:]:
                    print(f"    {line}")
        except subprocess.TimeoutExpired:
            print("  WARNING: Docket agent timed out after 1 hour")
        except Exception as e:
            print(f"  ERROR running docket agent: {e}")

    # ============================================================
    # TASK 3: Compare extracted trial dates with spreadsheet
    # ============================================================
    print(f"\n--- TASK 3: Comparing extracted trial dates with spreadsheet ---")

    # Load case data manager to get extracted trial dates
    sys.path.insert(0, os.path.join(os.getcwd(), "Scripts"))
    from case_data_manager import CaseDataManager
    dm = CaseDataManager()

    mismatches = []
    confirmed = []
    no_docket_data = []

    for row_idx, file_no, case_name, case_no, jurisdiction, spreadsheet_trial in to_run:
        extracted_trial = dm.get_value(file_no, "trial_date")

        if not extracted_trial:
            no_docket_data.append((file_no, case_name, jurisdiction))
            continue

        # Compare dates
        spreadsheet_date_str = ""
        if spreadsheet_trial:
            if isinstance(spreadsheet_trial, datetime.datetime):
                spreadsheet_date_str = spreadsheet_trial.strftime("%Y-%m-%d")
            elif isinstance(spreadsheet_trial, str):
                # Try to parse the formatted date back
                try:
                    d = datetime.datetime.strptime(spreadsheet_trial.strip(), "%B %d, %Y")
                    spreadsheet_date_str = d.strftime("%Y-%m-%d")
                except ValueError:
                    try:
                        d = datetime.datetime.strptime(spreadsheet_trial.strip(), "%Y-%m-%d")
                        spreadsheet_date_str = d.strftime("%Y-%m-%d")
                    except ValueError:
                        spreadsheet_date_str = str(spreadsheet_trial)

        extracted_str = str(extracted_trial).strip()

        if spreadsheet_date_str == extracted_str:
            confirmed.append((file_no, case_name, extracted_str))
        elif not spreadsheet_date_str and extracted_str:
            mismatches.append((file_no, case_name, "EMPTY in spreadsheet", extracted_str, row_idx))
        elif spreadsheet_date_str and not extracted_str:
            mismatches.append((file_no, case_name, spreadsheet_date_str, "NO TRIAL DATE in docket", row_idx))
        else:
            mismatches.append((file_no, case_name, spreadsheet_date_str, extracted_str, row_idx))

    # ============================================================
    # RESULTS
    # ============================================================
    print("\n" + "=" * 70)
    print("RESULTS")
    print("=" * 70)

    print(f"\n--- CONFIRMED TRIAL DATES ({len(confirmed)}) ---")
    for fn, name, date in confirmed:
        print(f"  {fn} ({name}): {date}")

    print(f"\n--- MISMATCHES ({len(mismatches)}) ---")
    for item in mismatches:
        fn, name, spreadsheet, docket = item[0], item[1], item[2], item[3]
        print(f"  {fn} ({name}): Spreadsheet={spreadsheet}, Docket={docket}")

    # Update spreadsheet with docket trial dates for mismatches
    wb2 = openpyxl.load_workbook(XLSX_PATH)
    ws2 = wb2.active
    updates_made = 0
    for item in mismatches:
        fn, name, spreadsheet_val, docket_val, ridx = item
        if docket_val and docket_val != "NO TRIAL DATE in docket":
            formatted = format_date_str(docket_val)
            if formatted:
                ws2.cell(row=ridx, column=7).value = formatted
                ws2.cell(row=ridx, column=7).number_format = '@'
                updates_made += 1
                print(f"  UPDATED {fn}: {spreadsheet_val} -> {formatted}")

    if updates_made > 0:
        wb2.save(XLSX_PATH)
        print(f"\n  Updated {updates_made} trial dates in spreadsheet based on docket data.")

    print(f"\n--- COULD NOT DOWNLOAD DOCKET ({len(cannot_run) + len(no_docket_data)}) ---")
    for fn, name, reason in cannot_run:
        print(f"  {fn} ({name}): {reason}")
    for fn, name, jurisdiction in no_docket_data:
        print(f"  {fn} ({name}): No trial date extracted from docket (jurisdiction: {jurisdiction})")

    print(f"\n--- DATE FORMATTING: Fixed {len(all_trial_date_rows)} dates to 'Month DD, YYYY' format ---")
    print("\nDone!")


if __name__ == "__main__":
    main()
