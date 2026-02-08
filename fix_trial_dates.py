import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import sqlite3
import subprocess
import time
import os

excel_path = r"C:\Users\ASerpik.DESKTOP-MRIMK0D\OneDrive - Bordin Semmer LLP\Desktop\AS Team Case List.xlsx"

# ============================================================
# Close Excel if file is open
# ============================================================
print("Closing Excel to release the file lock...")
# Use COM to tell Excel to save and close gracefully
close_script = r"""
$excel = [System.Runtime.InteropServices.Marshal]::GetActiveObject('Excel.Application')
if ($excel) {
    foreach ($wb in $excel.Workbooks) {
        if ($wb.Name -eq 'AS Team Case List.xlsx') {
            $wb.Save()
            $wb.Close($false)
            Write-Host 'Closed workbook'
        }
    }
}
"""
result = subprocess.run(
    ['powershell', '-Command', close_script],
    capture_output=True, text=True
)
print(f"  PowerShell output: {result.stdout.strip()}")
if result.stderr.strip():
    print(f"  PowerShell errors: {result.stderr.strip()}")
time.sleep(2)

# ============================================================
# 1. Fix the Excel file
# ============================================================
print(f"\nOpening Excel file: {excel_path}")

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

no_fill = PatternFill(fill_type=None)

fixes = {
    "5800.016": {
        "trial_date": datetime(2026, 9, 14),
        "changes": "Trial Date reverted - iCharlotte was wrong, docket confirms 09/14/2026",
    },
    "5800.038": {
        "trial_date": datetime(2027, 4, 12),
        "changes": "Trial Date reverted - iCharlotte was wrong, docket confirms 04/12/2027",
    },
}

found = {k: False for k in fixes}

for row in ws.iter_rows(min_row=2):
    cell_a = row[0]  # Column A = file number
    val = cell_a.value
    if val is None:
        continue

    # File numbers are floats in Excel; format to 3 decimal places for comparison
    try:
        file_str = f"{float(val):.3f}"
    except (ValueError, TypeError):
        continue

    if file_str in fixes:
        fix = fixes[file_str]
        row_num = cell_a.row

        # Column G = index 6 (Trial Date)
        cell_g = row[6]
        old_trial = cell_g.value
        cell_g.value = fix["trial_date"]

        # Remove yellow highlight from entire row
        for cell in row:
            if cell.fill and cell.fill.patternType:
                fg = cell.fill.fgColor
                if fg and fg.rgb:
                    rgb_str = str(fg.rgb).upper()
                    # Yellow variants: FFFFFF00, FFFF00, FFFFD700, etc.
                    if "FFFF" in rgb_str and rgb_str != "00000000":
                        cell.fill = no_fill

        # Column H = index 7 (Changes)
        cell_h = row[7]
        cell_h.value = fix["changes"]

        print(f"  Row {row_num}: {file_str} - Trial Date: {old_trial} -> {fix['trial_date'].date()}, highlight removed, Changes updated")
        found[file_str] = True

for k, v in found.items():
    if not v:
        print(f"  WARNING: File number {k} NOT FOUND in Excel!")

wb.save(excel_path)
print("Excel file saved successfully.\n")

# ============================================================
# Reopen in Excel
# ============================================================
print("Reopening file in Excel...")
os.startfile(excel_path)

# ============================================================
# 2. Fix the iCharlotte database
# ============================================================
db_path = r"C:\geminiterminal2\.gemini\case_data\master_cases.db"
print(f"\nConnecting to database: {db_path}")

conn = sqlite3.connect(db_path)
cur = conn.cursor()

cur.execute("UPDATE cases SET trial_date = '2026-09-14' WHERE file_number = '5800.016'")
rows_016 = cur.rowcount
print(f"  Updated 5800.016 trial_date to 2026-09-14 ({rows_016} row(s) affected)")

cur.execute("UPDATE cases SET trial_date = '2027-04-12' WHERE file_number = '5800.038'")
rows_038 = cur.rowcount
print(f"  Updated 5800.038 trial_date to 2027-04-12 ({rows_038} row(s) affected)")

conn.commit()

# Verify
cur.execute("SELECT file_number, case_name, trial_date FROM cases WHERE file_number IN ('5800.016', '5800.038')")
results = cur.fetchall()
print("\n  Verification:")
for r in results:
    print(f"    File: {r[0]}, Case: {r[1]}, Trial Date: {r[2]}")

conn.close()
print("\nDatabase updated and verified successfully.")
print("\n=== All fixes complete ===")
